import openpyxl
from utils.beta_excel_json_struct import read_excel_json
from openpyxl.styles import PatternFill, Font

import json
from utils.mapping_loader import derive_mapping

def normalize_sheet_name(name):
    return name.replace(" ", "").lower()

def apply_mapping_to_json(sheets):
    last_row_idx = 0
    mapping = derive_mapping()
    mapping_sheet = normalize_sheet_name(mapping["sheet_name"])
    for sheet in sheets:
        if normalize_sheet_name(sheet["sheet_name"]) == mapping_sheet:
            data = sheet["data"]
            styles = sheet.setdefault("styles", [])
            num_cols = len(data[0]) if data else 0

            #to add last_sr_no + 1 val (value for new serial number), we first need last_row_idx, which we get after checking y which is the place of append
           
            """
            in our logic, y is the row index where we want to insert the new data.
            We are doing so because values after y in sheets are performing totalling
            """
            # Find the row index y where data[y][3] == "C/F 2023-24"
            y = None
            for idx, row in enumerate(data):
                if len(row) > 3 and str(row[3]).strip() == "C/F 2023-24":
                    y = idx
                    break

            if y is not None:
                # Insert a blank row at y
                data.insert(y, [""] * num_cols)
                styles.insert(y, [{} for _ in range(num_cols)])
                target_row_idx = y
                last_row_idx = y
            else:
                # Fallback: append at the end as before
                last_row_idx = len(data) - 1
                while last_row_idx >= 0 and all((str(cell).strip() == "" for cell in data[last_row_idx])):
                    last_row_idx -= 1
                target_row_idx = last_row_idx + 1
                if target_row_idx >= len(data):
                    data.append([""] * num_cols)
                    styles.append([{} for _ in range(num_cols)])
                while len(styles) <= target_row_idx:
                    styles.append([{} for _ in range(num_cols)])

            # Sr. No. (auto-increment)
            last_sr_no = 0
            for i in range(1, last_row_idx + 1):
                try:
                    val = int(str(data[i][0]).strip())
                    if val > last_sr_no:
                        last_sr_no = val
                except Exception:
                    continue

            #serial number:
            data[target_row_idx][0] = str(last_sr_no + 1)
            
            # date_found
            data[target_row_idx][4] = mapping.get("date_found", "")

            # Listed/Unlisted and Rated/Unrated
            keywords = mapping.get("qualified_keywords", [])
            listed = next((k for k in keywords if k.lower() in ["listed", "unlisted"]), "")
            rated = next((k for k in keywords if k.lower() in ["rated", "unrated"]), "")
            data[target_row_idx][6] = f"{listed}/{rated}" if listed and rated else listed or rated

            # details
            details = mapping.get("details", {})
            data[target_row_idx][9] = str(details.get("units", ""))
            data[target_row_idx][10] = str(details.get("face_value", ""))
            data[target_row_idx][11] = str(details.get("issue_price", ""))
            data[target_row_idx][12] = str(details.get("nominal_value", ""))

                        # series
            data[target_row_idx][3] = mapping.get("series", "")

            # underlying_index
            data[target_row_idx][16] = mapping.get("underlying_value", "")

            # date_of_maturity
            data[target_row_idx][19] = mapping.get("date_of_maturity", "")

            # Highlight all these cells in green
            for col in [0, 3, 4, 6, 9, 10, 11, 12, 16, 19]:
                styles[target_row_idx][col] = {"fill": "FF00FF00", "font_color": "FF000000"}

            break  # Only apply to the first matching sheet
# In your run_json_excel_flow, call this before generate_xlsx_from_json:
def run_json_excel_flow():
    json_path = "configs/beta_excel.json"
    output_path = "output/generated_from_json.xlsx"

    sheets = read_excel_json(json_path)
    apply_mapping_to_json(sheets)  # <-- Add this line
    generate_xlsx_from_json(sheets, output_path)
    print("[✔] Excel generated from JSON:", output_path)

def apply_style(cell, style):
    fill = style.get("fill")
    if fill:
        if fill.startswith("theme:"):
            # openpyxl does not support setting theme colors directly, so skip or map as needed
            pass
        elif len(fill) == 8:  # ARGB
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")
        elif len(fill) == 6:  # RGB
            cell.fill = PatternFill(start_color="FF"+fill, end_color="FF"+fill, fill_type="solid")
        else:
            # fallback, try as is
            cell.fill = PatternFill(start_color=fill, end_color=fill, fill_type="solid")

    font_kwargs = {
        "bold": style.get("bold", False),
        "italic": style.get("italic", False)
    }
    font_color = style.get("font_color", None)
    if font_color:
        # Only use if it's a valid 8-char ARGB hex
        if isinstance(font_color, str) and len(font_color) == 8 and all(c in "0123456789ABCDEFabcdef" for c in font_color):
            font_kwargs["color"] = font_color
        # If it's 6-char RGB, prepend 'FF'
        elif isinstance(font_color, str) and len(font_color) == 6 and all(c in "0123456789ABCDEFabcdef" for c in font_color):
            font_kwargs["color"] = "FF" + font_color
        # Otherwise, skip setting color

    cell.font = Font(**font_kwargs)

def generate_xlsx_from_json(sheets, output_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet in sheets:
        ws = wb.create_sheet(title=sheet['sheet_name'])
        data = sheet['data']
        styles = sheet.get('styles', [])
        for i, row in enumerate(data):
            ws.append(row)
            if styles:
                for j, style in enumerate(styles[i]):
                    apply_style(ws.cell(row=i+1, column=j+1), style)
        # Re-apply merged cells if present
        if 'merged_cells' in sheet:
            for merge_range in sheet['merged_cells']:
                ws.merge_cells(merge_range)
    wb.save(output_path)

def run_json_excel_flow():
    json_path = "configs/beta_excel.json"
    output_path = "output/generated_from_json.xlsx"

    sheets = read_excel_json(json_path)
    apply_mapping_to_json(sheets)  # <-- Add this line
    generate_xlsx_from_json(sheets, output_path)
    print("[✔] Excel generated from JSON:", output_path)

if __name__ == "__main__":
    run_json_excel_flow()