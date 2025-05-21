from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import zipfile
from xml.etree import ElementTree

#sincel openpyxl doesnt read "theme" colors but cutie ms excel saves everything in theme colors :) we need to unzip xlsx to get theme_xml, and then use that to maintain the theme
#for some reason get_theme_colors is reading font color as white instead of black but the default color is always black, its reading it white meaning the color mapping is dogshit :)

def get_theme_colors(xlsx_path):
    with zipfile.ZipFile(xlsx_path, 'r') as archive:
        theme_xml = archive.read('xl/theme/theme1.xml')
        tree = ElementTree.fromstring(theme_xml)
        clr_scheme = tree.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}clrScheme')
        theme_colors = []
        for child in clr_scheme:
            srgb = child.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}srgbClr')
            sysclr = child.find('.//{http://schemas.openxmlformats.org/drawingml/2006/main}sysClr')
            if srgb is not None:
                theme_colors.append(srgb.attrib['val'].upper())
            elif sysclr is not None:
                theme_colors.append(sysclr.attrib['lastClr'].upper())
            else:
                theme_colors.append(None)
        return theme_colors

def get_rgb(color, theme_colors=None):
    if color is None:
        return None
    if hasattr(color, "rgb") and color.rgb is not None:
        rgb = color.rgb
        if isinstance(rgb, bytes):
            rgb = rgb.decode()
        if isinstance(rgb, str) and (len(rgb) == 8 or len(rgb) == 6):
            return rgb.upper()
    if hasattr(color, "theme") and color.theme is not None and theme_colors:
        idx = int(color.theme)
        if 0 <= idx < len(theme_colors):
            theme_rgb = theme_colors[idx]
            if theme_rgb:  # Only if not None
                return "FF" + theme_rgb
    return None

def get_cell_style(cell, theme_colors=None):
    style = {}
    # if cell.fill:
    #     print(f"cell.fill: {cell.fill}, patternType: {getattr(cell.fill, 'patternType', None)}")
    #     print(f"  fgColor: {vars(cell.fill.fgColor)}")
    #     print(f"  start_color: {vars(cell.fill.start_color)}")
    if cell.fill and isinstance(cell.fill, PatternFill):
        fg = get_rgb(cell.fill.fgColor, theme_colors)
        start = get_rgb(cell.fill.start_color, theme_colors)
        fill_color = start or fg
        print(f"  DEBUG: fg={fg}, start={start}, fill_color={fill_color}")
        if fill_color and fill_color != "00000000":
            style["fill"] = fill_color
    if cell.font:
        style["bold"] = cell.font.bold
        style["italic"] = cell.font.italic
        if cell.font.color:
            font_color = get_rgb(cell.font.color, theme_colors)
            # Convert white to black
            if font_color and font_color.upper() == "FFFFFFFF":
                font_color = "FF000000"
            # Only add if not default black or None
            if font_color and font_color != "FF000000":
                style["font_color"] = font_color
    return style

def excel_to_json_file(excel_path):
    theme_colors = get_theme_colors(excel_path)
    workbook = load_workbook(excel_path, data_only=True)
    result = {"sheets": []}
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        sheet_data = []
        style_data = []
        for row in ws.iter_rows():
            row_values = []
            row_styles = []
            for cell in row:
                row_values.append(str(cell.value).strip() if cell.value is not None else "")
                row_styles.append(get_cell_style(cell, theme_colors))
            if any(cell != "" for cell in row_values):
                sheet_data.append(row_values)
                style_data.append(row_styles)
        merged_cells = [str(merge) for merge in ws.merged_cells.ranges]
        result["sheets"].append({
            "sheet_name": sheet,
            "data": sheet_data,
            "styles": style_data,
            "merged_cells": merged_cells
        })
    return result